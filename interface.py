from openpyxl import *
from tkinter import *
import pandas as pd
import numpy as np
import testdata 
wb = load_workbook('C:/Users/SUTIRTHA PAL/Desktop/excel.xlsx')
sheets_dict = pd.ExcelFile('C:/Users/SUTIRTHA PAL/Desktop/excel.xlsx', sheet_name=None)
sheets_dict=sheets_dict.parse('Sheet1')
# create the sheet object
sheet = wb.active
def excel():

    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width=30
    sheet.column_dimensions['B'].width=10
    sheet.column_dimensions['C'].width=30
    sheet.column_dimensions['D'].width=10
    sheet.column_dimensions['E'].width=20
    sheet.column_dimensions['F'].width=40
    sheet.column_dimensions['G'].width=50
    sheet.column_dimensions['H'].width=20

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Username"
    sheet.cell(row=1, column=3).value = "Course"
    sheet.cell(row=1, column=4).value = "Semester"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"
    sheet.cell(row=1, column=8).value = "Password"

def register():
    global register_screen
    register_screen = Toplevel(main_screen)
    register_screen.title("Register")
    register_screen.geometry("900x600")
    excel()
    global username
    global password
    global name
    global email
    global contact
    global address
    global semester
    global course
    global username_entry
    global password_entry
    global name_entry
    global email_entry
    global contact_entry
    global address_entry
    global semester_entry
    global course_entry
    name= StringVar()
    username = StringVar()
    course = StringVar()
    semester = StringVar()
    contact= StringVar()
    email = StringVar()
    address = StringVar()
    password = StringVar()
    
    Label(register_screen, text="Please enter details below", bg="blue").pack()
    Label(register_screen, text="").pack()
    name_lable = Label(register_screen, text="Name * ")
    name_lable.pack()
    name_entry = Entry(register_screen, textvariable=name)
    name_entry.pack()
    username_lable = Label(register_screen, text="Username * ")
    username_lable.pack()
    username_entry = Entry(register_screen, textvariable=username)
    username_entry.pack()
    course_lable = Label(register_screen, text="Course * ")
    course_lable.pack()
    course_entry = Entry(register_screen, textvariable=course)
    course_entry.pack()
    semester_lable = Label(register_screen, text="Semester * ")
    semester_lable.pack()
    semester_entry = Entry(register_screen, textvariable=semester)
    semester_entry.pack()
    contact_lable = Label(register_screen, text="Contact * ")
    contact_lable.pack()
    contact_entry = Entry(register_screen, textvariable=contact)
    contact_entry.pack()
    email_lable = Label(register_screen, text="Email * ")
    email_lable.pack()
    email_entry = Entry(register_screen, textvariable=email)
    email_entry.pack()
    address_lable = Label(register_screen, text="Address * ")
    address_lable.pack()
    address_entry = Entry(register_screen, textvariable=address)
    address_entry.pack()
    password_lable = Label(register_screen, text="Password * ")
    password_lable.pack()
    password_entry = Entry(register_screen, textvariable=password, show='*')
    password_entry.pack()
    Label(register_screen, text="").pack()
    
    Button(register_screen, text="Register", width=10, height=1, bg="blue", command = register_user).pack()
 
 
# Designing window for login 
 
def login():
    global login_screen
    login_screen = Toplevel(main_screen)
    login_screen.title("Login")
    login_screen.geometry("500x350")
    Label(login_screen, text="Please enter details below to login").pack()
    Label(login_screen, text="").pack()
 
    global name_verify
    global username_verify
    global email_verify
    global address_verify
    global password_verify
 
    name_verify = StringVar()
    username_verify = StringVar()
    email_verify = StringVar()
    address_verify = StringVar()
    password_verify = StringVar()
    
    global name_login_entry
    global username_login_entry
    global email_login_entry
    global address_login_entry
    global password_login_entry
    
    Label(login_screen, text="Name * ").pack()
    name_login_entry = Entry(login_screen, textvariable=name_verify)
    name_login_entry.pack()
    Label(login_screen, text="").pack()
    Label(login_screen, text="Username * ").pack()
    username_login_entry = Entry(login_screen, textvariable=username_verify)
    username_login_entry.pack()
    Label(login_screen, text="").pack()
    Label(login_screen, text="Email * ").pack()
    email_login_entry = Entry(login_screen, textvariable=email_verify)
    email_login_entry.pack()
    Label(login_screen, text="").pack()
    Label(login_screen, text="Address * ").pack()
    address_login_entry = Entry(login_screen, textvariable=address_verify)
    address_login_entry.pack()
    Label(login_screen, text="").pack()
    Label(login_screen, text="Password * ").pack()
    password_login_entry = Entry(login_screen, textvariable=password_verify, show= '*')
    password_login_entry.pack()
    Label(login_screen, text="").pack()
    Button(login_screen, text="Login", width=10, height=1, command = login_verify).pack()
 
# Implementing event on register button
 
def register_user():
 
    excel()
    if (name_entry.get() == "" and
        username_entry.get() == "" and
        course_entry.get() == "" and
        semester_entry.get() == "" and
        contact_entry.get() == "" and
        email_entry.get() == "" and
        address_entry.get() == "" and
        password_entry.get()  == ""):

        print("empty input")

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = name_entry.get()
        sheet.cell(row=current_row + 1, column=2).value = username_entry.get()
        sheet.cell(row=current_row + 1, column=3).value = course_entry.get()
        sheet.cell(row=current_row + 1, column=4).value = semester_entry.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_entry.get()
        sheet.cell(row=current_row + 1, column=6).value = email_entry.get()
        sheet.cell(row=current_row + 1, column=7).value = address_entry.get()
        sheet.cell(row=current_row + 1, column=8).value = password_entry.get()

        # save the file
        wb.save('C:/Users/SUTIRTHA PAL/Desktop/excel.xlsx')

        # set focus on the name_field box
        #name_field.focus_set()

        # call the clear() function
        name_entry.delete(0,END)
        username_entry.delete(0,END)
        course_entry.delete(0,END)
        semester_entry.delete(0,END)
        contact_entry.delete(0,END)
        email_entry.delete(0,END)
        address_entry.delete(0,END)
        password_entry.delete(0,END)
 
    Label(register_screen, text="Registration Success", fg="green", font=("calibri", 11)).pack()
 
# Implementing event on login button 
 
def login_verify():
    username1 = username_verify.get()
    password1 = password_verify.get()
    username_login_entry.delete(0, END)
    password_login_entry.delete(0, END)
    li = sheets_dict.loc[sheets_dict.Username==username1,'Username']
    lii=li.unique()
    lis=sheets_dict.index[sheets_dict.Username==username1].tolist()
    if username1 in lii:
        file1 = sheets_dict.iloc[li.index[0]]['Password']
        if password1==file1:
            import featureextract
            if testdata.func(lis[0])==True:
               login_sucess()
            else:
               authentication_error()
 
        else:
            password_not_recognised()
 
    else:
        user_not_found()
 
# Designing popup for login success
 
def login_sucess():
    global login_success_screen
    login_success_screen = Toplevel(login_screen)
    login_success_screen.title("Success")
    login_success_screen.geometry("150x100")
    Label(login_success_screen, text="Login Success").pack()
    Button(login_success_screen, text="OK", command=delete_login_success).pack()
 
# Designing popup for login invalid password
 
def password_not_recognised():
    global password_not_recog_screen
    password_not_recog_screen = Toplevel(login_screen)
    password_not_recog_screen.title("Success")
    password_not_recog_screen.geometry("150x100")
    Label(password_not_recog_screen, text="Invalid Password ").pack()
    Button(password_not_recog_screen, text="OK", command=delete_password_not_recognised).pack()
 
# Designing popup for user not found
 
def user_not_found():
    global user_not_found_screen
    user_not_found_screen = Toplevel(login_screen)
    user_not_found_screen.title("Success")
    user_not_found_screen.geometry("150x100")
    Label(user_not_found_screen, text="User Not Found").pack()
    Button(user_not_found_screen, text="OK", command=delete_user_not_found_screen).pack()
    
def authentication_error():
    global error_screen
    error_screen=Toplevel(login_screen)
    error_screen.title("Success")
    error_screen.geometry("150x100")
    Label(error_screen,text="Authentication Error!!!Please try again").pack()
    Button(error_screen,text="OK",command=delete_error_screen).pack()
 
# Deleting popups

def delete_error_screen():
    error_screen.destroy()
def delete_login_success():
    login_success_screen.destroy()
 
 
def delete_password_not_recognised():
    password_not_recog_screen.destroy()
 
 
def delete_user_not_found_screen():
    user_not_found_screen.destroy()
 
 
# Designing Main(first) window
 
def main_account_screen():
    global main_screen
    main_screen = Tk()
    main_screen.geometry("300x250")
    main_screen.title("Account Login")
    Label(text="Select Your Choice", bg="blue", width="300", height="2", font=("Calibri", 13)).pack()
    Label(text="").pack()
    Button(text="Login", height="2", width="30", command = login).pack()
    Label(text="").pack()
    Button(text="Register", height="2", width="30", command=register).pack()
 
    main_screen.mainloop()
 
if __name__ == "__main__":
    main_account_screen()
